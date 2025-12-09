"""
Seating Arrangement Allocation System
--------------------------------------
This system allocates students to exam halls with the following features:
1. Alternating department seating (different departments seated alternately)
2. Supports arrear exam takers from different years
3. Generates Excel output with two formats:
   - Format 1: Same department students together
   - Format 2: Different department students alternating
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os


class SeatingAllocationSystem:
    def __init__(self, halls_file, students_file):
        """Initialize the seating allocation system"""
        # Read halls data
        self.halls_df = pd.read_csv(halls_file)
        self.halls_df.columns = self.halls_df.columns.str.strip()
        
        # Read students data - preserve register numbers as strings
        self.students_df = pd.read_csv(students_file, dtype={'Register Number': str})
        self.students_df.columns = self.students_df.columns.str.strip()
        
        # Prepare data structures
        self.allocations = []
        self.hall_wise_allocations = {}
        
    def allocate_seats_mixed_department(self):
        """
        Allocate seats linearly by department
        Complete one department before moving to the next
        """
        print("=" * 60)
        print("SEATING ALLOCATION - LINEAR DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Sort students by department and register number
        students_sorted = self.students_df.sort_values(
            ['Department', 'Register Number']
        ).reset_index(drop=True)
        
        # Allocate students to halls
        current_hall_idx = 0
        current_seat_in_hall = 1
        
        allocations = []
        
        for idx, student in students_sorted.iterrows():
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            allocations.append({
                'Hall No': hall_no,
                'Seat No': current_seat_in_hall,
                'Register Number': student['Register Number'],
                'Name': student['Name'],
                'Department': student['Department']
            })
            
            current_seat_in_hall += 1
            
            # Check if we need to move to next hall
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
                
                if current_hall_idx >= len(self.halls_df):
                    print("Warning: Ran out of halls!")
                    break
        
        self.allocations = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {len(self.allocations)}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        # Create hall-wise summary
        self._create_hall_wise_summary()
        
        return self.allocations
    
    def allocate_seats_alternating_department(self):
        """
        Allocate seats with different departments alternating
        This ensures students from the same department don't sit next to each other
        """
        print("\n" + "=" * 60)
        print("SEATING ALLOCATION - ALTERNATING DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Group students by department
        departments = self.students_df['Department'].unique()
        dept_groups = {dept: self.students_df[self.students_df['Department'] == dept].copy() 
                      for dept in departments}
        
        # Sort each department group by register number
        for dept in dept_groups:
            dept_groups[dept] = dept_groups[dept].sort_values('Register Number').reset_index(drop=True)
        
        # Create pointers for each department
        dept_pointers = {dept: 0 for dept in departments}
        dept_list = list(departments)
        
        # Allocate students to halls
        current_hall_idx = 0
        current_seat_in_hall = 1
        current_dept_idx = 0
        
        allocations = []
        
        total_allocated = 0
        total_students = len(self.students_df)
        
        while total_allocated < total_students:
            # Get current hall info
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            # Try to allocate from current department
            dept = dept_list[current_dept_idx]
            dept_pointer = dept_pointers[dept]
            
            # Check if current department still has students
            if dept_pointer < len(dept_groups[dept]):
                student = dept_groups[dept].iloc[dept_pointer]
                
                allocations.append({
                    'Hall No': hall_no,
                    'Seat No': current_seat_in_hall,
                    'Register Number': student['Register Number'],
                    'Name': student['Name'],
                    'Department': student['Department']
                })
                
                dept_pointers[dept] += 1
                current_seat_in_hall += 1
                total_allocated += 1
                
            # Move to next department in rotation
            current_dept_idx = (current_dept_idx + 1) % len(dept_list)
            
            # Check if we need to move to next hall
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
                
                if current_hall_idx >= len(self.halls_df):
                    print("Warning: Ran out of halls!")
                    break
        
        allocations_df = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {total_allocated}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        return allocations_df
    
    def allocate_seats_same_department(self):
        """
        Allocate seats with same department students together
        Students from same department fill up halls before moving to next department
        """
        print("\n" + "=" * 60)
        print("SEATING ALLOCATION - SAME DEPARTMENT FORMAT")
        print("=" * 60)
        
        # Sort students by department and register number
        students_sorted = self.students_df.sort_values(
            ['Department', 'Register Number']
        ).reset_index(drop=True)
        
        allocations = []
        current_hall_idx = 0
        current_seat_in_hall = 1
        
        for idx, student in students_sorted.iterrows():
            hall_no = self.halls_df.loc[current_hall_idx, 'hallno']
            hall_capacity = self.halls_df.loc[current_hall_idx, 'capacity']
            
            allocations.append({
                'Hall No': hall_no,
                'Seat No': current_seat_in_hall,
                'Register Number': student['Register Number'],
                'Name': student['Name'],
                'Department': student['Department']
            })
            
            current_seat_in_hall += 1
            
            # Move to next hall if current hall is full
            if current_seat_in_hall > hall_capacity:
                current_hall_idx += 1
                current_seat_in_hall = 1
        
        allocations_df = pd.DataFrame(allocations)
        print(f"\nTotal students allocated: {len(allocations_df)}")
        print(f"Halls used: {current_hall_idx + 1} out of {len(self.halls_df)}")
        
        return allocations_df
    
    def _create_hall_wise_summary(self):
        """Create a summary of allocations by hall"""
        self.hall_wise_allocations = {}
        
        for hall_no in self.allocations['Hall No'].unique():
            hall_data = self.allocations[self.allocations['Hall No'] == hall_no].copy()
            hall_data = hall_data.sort_values('Seat No').reset_index(drop=True)
            self.hall_wise_allocations[hall_no] = hall_data
    
    def generate_excel_report(self, output_file='seating_allocation_report.xlsx'):
        """Generate comprehensive Excel report with multiple sheets"""
        
        print("\n" + "=" * 60)
        print("GENERATING EXCEL REPORT")
        print("=" * 60)
        
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        
        # Sheet 1: Complete allocation list (Linear Department)
        self.allocations.to_excel(writer, sheet_name='Complete Allocation', index=False)
        
        # Sheet 2: Hall-wise breakdown
        hall_summary = []
        for hall_no, hall_data in sorted(self.hall_wise_allocations.items()):
            dept_counts = hall_data['Department'].value_counts()
            hall_summary.append({
                'Hall No': hall_no,
                'Total Students': len(hall_data),
                'Departments': ', '.join([f"{dept}({count})" for dept, count in dept_counts.items()])
            })
        
        pd.DataFrame(hall_summary).to_excel(writer, sheet_name='Hall Summary', index=False)
        
        # Sheet 3: Department-wise summary
        dept_summary = self.allocations.groupby('Department').agg({
            'Register Number': 'count',
            'Hall No': lambda x: f"{x.min()} to {x.max()}"
        }).reset_index()
        dept_summary.columns = ['Department', 'Total Students', 'Hall Range']
        dept_summary.to_excel(writer, sheet_name='Department Summary', index=False)
        
        # Create individual hall sheets
        for hall_no, hall_data in sorted(self.hall_wise_allocations.items()):
            sheet_name = f"Hall {hall_no}"
            hall_data.to_excel(writer, sheet_name=sheet_name, index=False)
        
        writer.close()
        
        # Now format the Excel file
        self._format_excel(output_file)
        
        print(f"\n✓ Excel report generated: {output_file}")
        print(f"✓ Total sheets created: {3 + len(self.hall_wise_allocations)}")
        
        return output_file
    
    def _format_excel(self, file_path):
        """Apply formatting to the Excel file"""
        from openpyxl import load_workbook
        from openpyxl.styles import numbers
        
        wb = load_workbook(file_path)
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find the Register Number column
            reg_num_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Register Number':
                    reg_num_col = idx
                    break
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Format data cells
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for idx, cell in enumerate(row, 1):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Force Register Number column to be text format
                    if reg_num_col and idx == reg_num_col:
                        cell.number_format = '@'  # Text format
                        # Ensure the value is stored as string
                        if cell.value is not None:
                            cell.value = str(cell.value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(file_path)
    
    def print_statistics(self):
        """Print allocation statistics"""
        print("\n" + "=" * 60)
        print("ALLOCATION STATISTICS")
        print("=" * 60)
        
        print("\nDepartment-wise allocation:")
        dept_stats = self.allocations.groupby('Department').agg({
            'Register Number': 'count',
            'Hall No': ['min', 'max']
        })
        
        for dept in dept_stats.index:
            count = dept_stats.loc[dept, ('Register Number', 'count')]
            hall_min = dept_stats.loc[dept, ('Hall No', 'min')]
            hall_max = dept_stats.loc[dept, ('Hall No', 'max')]
            print(f"  {dept:8s}: {count:3d} students (Halls {hall_min:2d} to {hall_max:2d})")
        
        print("\nHall utilization:")
        for hall_no in sorted(self.allocations['Hall No'].unique()):
            hall_capacity = self.halls_df[self.halls_df['hallno'] == hall_no]['capacity'].values[0]
            allocated = len(self.allocations[self.allocations['Hall No'] == hall_no])
            utilization = (allocated / hall_capacity) * 100
            print(f"  Hall {hall_no:2d}: {allocated:2d}/{hall_capacity:2d} seats ({utilization:5.1f}% utilized)")


def main():
    """Main execution function"""
    print("\n" + "=" * 60)
    print("SEATING ARRANGEMENT ALLOCATION SYSTEM")
    print("First Year Students - Academic Year 2024-25")
    print("=" * 60)
    
    # File paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    halls_file = os.path.join(script_dir, 'halls.csv')
    students_file = os.path.join(script_dir, 'year1.csv')
    output_file = os.path.join(script_dir, 'first_year_seating_allocation.xlsx')
    
    # Create allocation system
    system = SeatingAllocationSystem(halls_file, students_file)
    
    # Perform allocation (Linear Department Format - main allocation)
    allocations = system.allocate_seats_mixed_department()
    
    # Generate comprehensive Excel report
    system.generate_excel_report(output_file)
    
    # Print statistics
    system.print_statistics()
    
    print("\n" + "=" * 60)
    print("ALLOCATION COMPLETE!")
    print("=" * 60)
    print(f"\nOutput file: {output_file}")
    print("\nThe Excel file contains:")
    print("  1. Complete Allocation - Linear department format")
    print("  2. Hall Summary - Overview of each hall")
    print("  3. Department Summary - Department-wise statistics")
    print("  4. Individual Hall Sheets - Detailed view for each hall")
    print("\n")


if __name__ == "__main__":
    main()
